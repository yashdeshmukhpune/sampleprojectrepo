import openpyxl

Workbook=openpyxl.load_workbook("C:/Users/devya/PycharmProjects/pythonProjectdemopytest/testData/file_example_XLSX_50.xlsx")
testsheet=Workbook["Sheet1"]
print(testsheet.max_row)
print(testsheet.max_column)
print(testsheet.cell(2,2).value)
